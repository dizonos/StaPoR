import openpyxl
from openpyxl import Workbook
import sys
import matplotlib.pyplot as plt
from data import db_session
import docx
import os
import csv
import shutil
from PyQt5 import uic
from cryptography.fernet import Fernet
from PyQt5.QtWidgets import QApplication, QMainWindow, QDialog, QTableWidgetItem, QWidget
from PyQt5.QtWidgets import QFileDialog, QMessageBox, QInputDialog, QPlainTextEdit
from data.pupils import Pupil
from data.form import Class
from data.journal import Journal
from data.work import Work


def mistake(warning, thing_to_delete=None):
    msg = QMessageBox()
    msg.setIcon(QMessageBox.Critical)
    msg.setText("Error")
    msg.setInformativeText(warning)
    msg.setWindowTitle("Error")
    msg.exec_()
    if thing_to_delete:
        db_sess.delete(thing_to_delete)
        db_sess.commit()
    return


def write_key():
    key = Fernet.generate_key()
    return key


def encrypt(filename, key):
    f = Fernet(key)
    if filename.split('.')[-1] == 'txt':
        with open(filename, 'rb') as file:
            file_data = file.read()
    else:
        doc = docx.Document(filename)
        all_paras = doc.paragraphs
        file_data = '\n'.join(i.text for i in all_paras).encode('utf-8')
        os.remove(filename)
        filename = filename.split('.')[0] + '.txt'
    encrypted_data = f.encrypt(file_data)
    with open(f'{filename}', 'wb') as file:
        file.write(encrypted_data)
    file.close()


def decrypt(filename, key):
    f = Fernet(key)
    if filename.split('.')[-1] == 'txt':
        with open(filename, 'rb') as file:
            encrypted_data = file.read()
        decrypted_data = f.decrypt(encrypted_data)
        file.close()
    return decrypted_data.decode('utf-8')


def update_journal():
    lines = db_sess.query(Journal).all()
    for i in range(len(lines)):
        line = lines[i]
        line.id = i + 1
        db_sess.commit()


def update_pupils():
    lines = db_sess.query(Pupil).all()
    for i in range(len(lines)):
        line = lines[i]
        line.id = i + 1
        db_sess.commit()


def update_works():
    lines = db_sess.query(Work).all()
    for i in range(len(lines)):
        line = lines[i]
        line.id = i + 1
        db_sess.commit()


def update_form():
    lines = db_sess.query(Class).all()
    for i in range(len(lines)):
        line = lines[i]
        line.id = i + 1
        db_sess.commit()


class ShowWork(QWidget):
    def __init__(self, res, name, work):
        super().__init__()
        self.work = work
        self.title = name
        self.initUI(res)

    def initUI(self, res):
        self.setGeometry(300, 300, 700, 600)
        self.setWindowTitle(self.title)
        self.text = QPlainTextEdit(self)
        self.text.resize(661, 481)
        self.text.move(20, 10)
        self.text.setPlainText(res)


class MainForm(QMainWindow):
    def __init__(self):
        super().__init__()
        uic.loadUi('untitled.ui', self)
        db_sess = db_session.create_session()
        forms = db_sess.query(Class).all()
        for i in forms:
            self.comboBox_3.addItem(i.form)
        classes = db_sess.query(Class).filter(Class.form == self.comboBox_3.currentText()).first()
        if classes:
            pupils = db_sess.query(Pupil).filter(Pupil.form == classes.id).all()
            for i in pupils:
                self.comboBox.addItem(i.full_name)
        self.comboBox_3.currentIndexChanged.connect(self.main_table)
        self.pushButton.clicked.connect(self.create_table)
        self.pushButton_2.clicked.connect(self.open_table)
        self.pushButton_3.clicked.connect(self.open_without_interface)
        self.pushButton_8.clicked.connect(self.del_person)
        self.pushButton_9.clicked.connect(self.add_person)
        self.pushButton_6.clicked.connect(self.add_class)
        self.pushButton_11.clicked.connect(self.del_class)
        self.pushButton_4.clicked.connect(self.del_table)
        self.pushButton_5.clicked.connect(self.diagrams)
        self.pushButton_10.clicked.connect(self.export)
        self.pushButton_7.clicked.connect(self.change)
        self.pushButton_12.clicked.connect(self.close)
        self.main_table()

    def show_work(self, name):
        work = db_sess.query(Work).filter(Work.title == name).first()
        self.file_name = work.file_name
        self.key = work.key
        if not self.file_name or not self.key:
            return -1
        res = decrypt(self.file_name, self.key)
        self.form2 = ShowWork(res, work.title, work)
        return self.form2

    def export(self):
        file = QFileDialog.getSaveFileName(self, 'Save File', '', 'Таблица excel (*.xlsx);;Таблица csv (*.csv)')
        if not file[0]:
            return
        cl = db_sess.query(Class).filter(Class.form == self.comboBox_3.currentText()).first().id
        spisok = db_sess.query(Work).filter(cl == Work.form).all()
        spisok = [i.title for i in spisok]
        spisok.insert(0, 'Журнал')
        name, ok_pressed = QInputDialog.getItem(
            self, "Выберите таблицу", "Какую таблицу экспортировать?",
            spisok, 1, False)
        if ok_pressed:
            pupils = db_sess.query(Pupil).filter(cl == Pupil.form).all()
            pupils = {i.id: i.full_name for i in pupils}
            works = db_sess.query(Work).filter(cl == Work.form).all()
            works = {i.id: i.title for i in works}
            if file[1] == 'Таблица excel (*.xlsx)':
                wb = Workbook(write_only=True)
                ws = wb.create_sheet(name)
                if name == 'Журнал':
                    header = ['№', 'ФИО', 'Название работы', 'Оценка', 'Вариант', 'Баллы за задания']
                    ws.append(header)
                    journal = db_sess.query(Journal).filter(cl == Journal.pupil_form).all()
                    for i in range(len(journal)):
                        j = journal[i]
                        line = [i + 1, pupils[j.full_name], works[j.task_name], j.mark, j.version, j.score_for_task]
                        ws.append(line)
                else:
                    works = db_sess.query(Work).filter(cl == Work.form, Work.title == name).first()
                    journal = db_sess.query(Journal).filter(cl == Journal.pupil_form, Journal.task_name == works.id).all()
                    header = ['№', 'ФИО', 'Класс', 'Оценка', 'Вариант']
                    for i in range(len(journal[0].score_for_task.split(';'))):
                        header.append(f'Задание №{i + 1}')
                    ws.append(header)
                    for i in range(len(journal)):
                        j = journal[i]
                        line = [i + 1, pupils[j.full_name], self.comboBox_3.currentText(), j.mark, j.version]
                        sup = [int(i) for i in j.score_for_task.split(';')]
                        line += sup
                        ws.append(line)
                wb.save(file[0])
            else:
                path = file[0] + '.csv'
                with open(path, 'w', newline='', encoding='utf-8') as file:
                    filewriter = csv.writer(file, delimiter=';')
                    if name == 'Журнал':
                        header = ['№', 'ФИО', 'Название работы', 'Оценка', 'Вариант', 'Баллы за задания']
                        filewriter.writerow(header)
                        journal = db_sess.query(Journal).filter(cl == Journal.pupil_form).all()
                        for i in range(len(journal)):
                            j = journal[i]
                            line = [i + 1, pupils[j.full_name], works[j.task_name], j.mark, j.version, j.score_for_task]
                            filewriter.writerow(line)
                    else:
                        works = db_sess.query(Work).filter(cl == Work.form, Work.title == name).first()
                        journal = db_sess.query(Journal).filter(cl == Journal.pupil_form, Journal.task_name == works.id).all()
                        header = ['№', 'ФИО', 'Класс', 'Оценка', 'Вариант']
                        for i in range(len(journal[0].score_for_task.split(';'))):
                            header.append(f'Задание №{i + 1}')
                        filewriter.writerow(header)
                        for i in range(len(journal)):
                            j = journal[i]
                            line = [i + 1, pupils[j.full_name], self.comboBox_3.currentText(), j.mark, j.version]
                            sup = [int(i) for i in j.score_for_task.split(';')]
                            line += sup
                            filewriter.writerow(line)
                    file.close()

    def diagrams(self):
        name = self.comboBox.currentText()
        name = db_sess.query(Pupil).filter(Pupil.full_name == name).first().id
        if self.comboBox_2.currentText() == 'Успеваемость ученика':
            lines = db_sess.query(Journal).filter(Journal.full_name == name).all()
            marks = [i.mark for i in lines]
            tasks = ['\n'.join(j for j in db_sess.query(Work).filter(Work.id == i.task_name).first().title.split()) for i in lines]
            print(tasks)
            plt.axis([0, 5, 0, len(tasks)])
            plt.barh(tasks, marks)
            plt.xlabel('Оценка за работу')
            plt.ylabel('Название работы')
        elif self.comboBox_2.currentText() == 'Количество пропусков':
            classes = db_sess.query(Class).filter(Class.form == self.comboBox_3.currentText()).first().id
            all = db_sess.query(Work).filter(Work.form == classes).all()
            wasnt = 0
            for i in all:
                mark = db_sess.query(Journal).filter(Journal.task_name == i.id, Journal.full_name == name).first().mark
                if mark == -1:
                    wasnt += 1
            all = len(all)
            print(all, wasnt)
            plt.pie([all - wasnt, wasnt], autopct='%.2f', labels=['Был', "Не был"])
        elif self.comboBox_2.currentText() == 'Успеваемость за работу':
            classes = db_sess.query(Class).filter(Class.form == self.comboBox_3.currentText()).first().id
            works = db_sess.query(Work).filter(Work.form == classes).all()
            works_with_no_n = list()
            for i in works:
                mark = db_sess.query(Journal).filter(Journal.task_name == i.id, Journal.full_name == name).first().id
                if mark != -1:
                    works_with_no_n.append(i.title)
            task_name, ok_pressed = QInputDialog.getItem(self, 'Выбор работы', 'Выберите работу', works_with_no_n)
            form2 = self.show_work(task_name)
            work_id = db_sess.query(Work).filter(Work.title == task_name).first().id
            if not ok_pressed:
                return
            works = db_sess.query(Journal).filter(Journal.pupil_form == classes, Journal.task_name == work_id).all()
            marks_names = dict()
            for i in works:
                name = db_sess.query(Pupil).filter(Pupil.id == i.full_name).first()
                mark = i.mark
                if mark != -1:
                    if mark not in marks_names.keys():
                        marks_names[mark] = [name.full_name + f' ({mark})']
                    else:
                        marks_names[mark].append(name.full_name + f' ({mark})')
            marks = list()
            names = list()
            for i in marks_names.keys():
                marks.append(len(marks_names[i]))
                names.append('\n'.join(i for i in marks_names[i]))
            plt.pie(marks, labels=names, autopct='%.2f')
        elif self.comboBox_2.currentText() == 'Баллы за задания':
            classes = db_sess.query(Class).filter(Class.form == self.comboBox_3.currentText()).first().id
            works = db_sess.query(Work).filter(Work.form == classes).all()
            works_with_no_n = list()
            for i in works:
                mark = db_sess.query(Journal).filter(Journal.task_name == i.id, Journal.full_name == name).first().id
                if mark != -1:
                    works_with_no_n.append(i.title)
            task_name, ok_pressed = QInputDialog.getItem(self, 'Выбор работы', 'Выберите работу', works_with_no_n)
            form2 = self.show_work(task_name)
            if not ok_pressed:
                return
            work_id = db_sess.query(Work).filter(Work.title == task_name).first().id
            work = db_sess.query(Journal).filter(Journal.full_name == name, Journal.task_name == work_id, Journal.pupil_form == classes).first()
            scores_p = [int(i) for i in work.score_for_task.split(';')]
            works = db_sess.query(Journal).filter(Journal.task_name == work_id, Journal.pupil_form == classes).all()
            scores_a = list()
            for i in works:
                sp = [int(j) for j in i.score_for_task.split(';')]
                scores_a += sp
            persent_score = [scores_p[i] / max(scores_a[i::len(scores_p)]) for i in range(len(scores_p))]
            tasks_name = [f'Задание {i}' for i in range(1, len(scores_p) + 1)]
            plt.axis([1, len(tasks_name), 0, 1])
            plt.ylabel('Отношение балла ученика\nк максимальному(по классу)')
            plt.xlabel('Задания')
            plt.bar(tasks_name, persent_score)
        if (self.comboBox_2.currentText() == 'Успеваемость за работу' or
            self.comboBox_2.currentText() == 'Баллы за задания') and form2 != -1:
            form2.show()
        plt.show()
        msgBox = QMessageBox()
        msgBox.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
        msgBox.setText("Создать отчёт?")
        result = msgBox.exec_()
        if QMessageBox.No == result:
            return
        file = QFileDialog.getSaveFileName(self, 'Save File', '', 'Файл c отчётом(*.txt)')
        if not file[0] or self.comboBox_2.currentText() == 'Количество пропусков':
            return
        with open(file[0], 'w', encoding='utf-8') as f:
            if self.comboBox_2.currentText() == 'Успеваемость ученика':
                f.write(f'Отчет по успеваемости {self.comboBox.currentText()}\n')
                tasks = [db_sess.query(Work).filter(Work.id == i.task_name).first().title for i in lines]
                for i in range(len(lines)):
                    f.write(f'Работа ({tasks[i]}) была написана на {marks[i]}\n')
                marks_works = [(marks[i], tasks[i]) for i in range(len(lines))]
                worst = [i for i in marks_works if i[0] == min(marks) and i[0] != max(marks)]
                if worst:
                    f.write("Хуже всего написаны работы:\n")
                    for i in worst:
                        f.write(f'{i[1]} на оценку {i[0]}\n')
                best = [i for i in marks_works if i[0] == max(marks) and i[0] != min(marks)]
                if best:
                    f.write("Лучше всего написаны работы:\n")
                    for i in best:
                        f.write(f'{i[1]} на оценку {i[0]}\n')
            elif self.comboBox_2.currentText() == 'Успеваемость за работу':
                f.write(f'Успеваемость по работе {task_name}\n')
                tup_works = [(i, marks_names[i]) for i in marks_names.keys()]
                tup_works.sort(key=lambda x: x[0])
                for i in tup_works:
                    f.write(f'{i[1]} написал/а работу на {i[0]}\n')
                w = tup_works[0][0]
                m = tup_works[::-1][0][0]
                worst = [i for i in tup_works if i[0] == w and i[0] != m]
                if worst:
                    f.write(f'Минимальный балл ({w}) за работу получил')
                    f.write('и\n' if len(worst) > 1 else '/а\n')
                    f.write('\n')
                    for i in worst:
                        f.write(f'i{1}\n')
                best = [i for i in tup_works if i[0] == m and i[0] != w]
                if best:
                    f.write(f'Максимальный балл ({m}) за работу получил')
                    f.write('и\n' if len(worst) > 1 else '/а\n')
                    f.write('\n')
                    for i in best:
                        f.write(f'{i[1]}\n')
            elif self.comboBox_2.currentText() == 'Баллы за задания':
                f.write(f'Баллы {self.comboBox.currentText()} за работу: {task_name}\n')
                scores_names = dict()
                for i in range(len(tasks_name)):
                    if scores_p[i] not in scores_names.keys():
                        scores_names[scores_p[i]] = [tasks_name[i]]
                    else:
                        scores_names[scores_p[i]].append(tasks_name[i])
                tup_works = [(i, scores_names[i]) for i in scores_names.keys()]
                tup_works.sort(key=lambda x: x[0])
                for i in tup_works:
                    tasks = ", ".join(_ for _ in i[1])
                    f.write(f'За {tasks} получил/а {i[0]}\n')

    def del_table(self):
        name, ok_pressed = QInputDialog.getText(self, "Удаление работы",
                                                "Введите название работы")
        if ok_pressed:
            classes = db_sess.query(Class).filter(Class.form == self.comboBox_3.currentText()).first()
            work = db_sess.query(Work).filter(Work.title == name, Work.form == classes.id).first()
            if not work:
                mistake('Такой работы не существует')
                return
            lines = db_sess.query(Journal).filter(Journal.task_name == work.id, Journal.pupil_form == classes.id).all()
            for i in lines:
                db_sess.delete(i)
                db_sess.commit()
            n = work.id
            db_sess.delete(work)
            db_sess.commit()
            lines = db_sess.query(Journal).filter(Journal.task_name > n, Journal.pupil_form == classes.id).all()
            for i in lines:
                i.task_name -= 1
                db_sess.commit()
            update_journal()
            update_works()
            self.main_table()

    def add_class(self):
        name, ok_pressed = QInputDialog.getText(self, 'Добавить класс', "Введите название класса")
        if ok_pressed:
            form = Class(
                form=name
            )
            db_sess.add(form)
            db_sess.commit()
            self.comboBox_3.addItem(name)
        self.main_table()

    def del_class(self):
        name, ok_pressed = QInputDialog.getText(self, 'Удалить класс', "Введите название класса")
        if ok_pressed:
            form = db_sess.query(Class).filter(Class.form == name).first()
            if form:
                journal = db_sess.query(Journal).filter(Journal.pupil_form == form.id).all()
                for i in journal:
                    db_sess.delete(i)
                    db_sess.commit()
                pupils = db_sess.query(Pupil).filter(Pupil.form == form.id).all()
                for i in pupils:
                    db_sess.delete(i)
                    db_sess.commit()
                works = db_sess.query(Work).filter(Work.form == form.id).all()
                for i in works:
                    db_sess.delete(i)
                    db_sess.commit()
                journal = db_sess.query(Journal).filter(Journal.pupil_form > form.id).all()
                for i in journal:
                    i.pupil_form -= 1
                    db_sess.commit()
                pupils = db_sess.query(Pupil).filter(Pupil.form > form.id).all()
                for i in pupils:
                    i.form -= 1
                    db_sess.commit()
                works = db_sess.query(Work).filter(Work.form == form.id).all()
                for i in works:
                    i.form -= 1
                    db_sess.commit()
                n = form.id - 1
                self.comboBox_3.removeItem(n)
                db_sess.delete(form)
                db_sess.commit()
                update_journal()
                update_form()
                update_pupils()
        self.main_table()

    def add_person(self):
        name, ok_pressed = QInputDialog.getText(self, "Добавить ученика",
                                                 "Введите ФИО ученика")
        if ok_pressed:
            print(name)
            if not all(x.isalpha() or x.isspace() for x in name):
                mistake(f'Неверно написано имя ученика')
                return
            if len(name.split()) < 2:
                mistake('У ученика должны быть имя и фамилия')
                return
            classes = db_sess.query(Class).filter(Class.form == self.comboBox_3.currentText()).first()
            person = Pupil(
                full_name=name,
                form=classes.id
            )
            db_sess.add(person)
            db_sess.commit()
            works = db_sess.query(Work).all()
            for i in works:
                line = Journal(
                    full_name=person.id,
                    pupil_form=person.form,
                    task_name=i.id,
                    version=1,
                    mark=-1
                )
                db_sess.add(line)
                db_sess.commit()

        self.main_table()

    def del_person(self):
        name, ok_pressed = QInputDialog.getText(self, "Удалить ученика", "Введите ФИО ученика")
        if ok_pressed:
            classes = db_sess.query(Class).filter(Class.form == self.comboBox_3.currentText()).first()
            person = db_sess.query(Pupil).filter(Pupil.full_name == name).first()
            if not person:
                mistake('Такого ученика не существует')
                return
            works = db_sess.query(Journal).filter(Journal.full_name == person.id).all()
            for i in works:
                db_sess.delete(i)
                db_sess.commit()
            n = person.id - 1
            self.comboBox.removeItem(n)
            db_sess.delete(person)
            db_sess.commit()
            works = db_sess.query(Journal).filter(Journal.full_name > person.id, Journal.pupil_form == classes.id).all()
            for i in works:
                i.full_name -= 1
                db_sess.commit()
        update_pupils()
        update_journal()
        self.main_table()

    def open_table(self):
        fname = QFileDialog.getOpenFileName(
            self, 'Выбрать таблицу', '',
            'Таблица (*.xlsx);')[0]
        if not fname:
            return
        self.form1 = AddTable(fname)
        self.form1.show()

    def create_table(self):
        self.form1 = NewTable('Создание таблицы с результатами')
        self.form1.show()
        self.main_table()

    def open_without_interface(self):
        fname = QFileDialog.getOpenFileName(
            self, 'Выбрать таблицу', '',
            'Таблица (*.xlsx);')[0]
        if not fname:
            return
        self.form1 = AddTable(fname)
        self.form1.progress.hide()
        self.form1.save_table()
        self.main_table()

    def change(self):
        cl = db_sess.query(Class).filter(Class.form == self.comboBox_3.currentText()).first().id
        spisok = db_sess.query(Work).filter(cl == Work.form).all()
        spisok = [i.title for i in spisok]
        name, ok_pressed = QInputDialog.getItem(
            self, "Выберите таблицу", "Какую таблицу экспортировать?",
            spisok, 1, False)
        if ok_pressed:
            self.form1 = NewTable('Изменение результатов работы')
            self.form1.pushButton_3.hide()
            self.form1.change_work(name)
            self.form1.show()
            update_journal()


    def main_table(self):
        classes = db_sess.query(Class).filter(Class.form == self.comboBox_3.currentText()).first()
        if not classes:
            return
        work_titles = db_sess.query(Work).filter(Work.form == classes.id).all()
        work_titles = [i.title for i in work_titles]
        header = list()
        header.append(('ФИО'))
        header += work_titles
        names = db_sess.query(Pupil).filter(Pupil.form == classes.id).all()
        names = [i.full_name for i in names]
        names.sort(key=lambda x: x.split()[1])
        self.tableWidget.clear()
        self.tableWidget.setColumnCount(len(header))
        self.tableWidget.setRowCount(0)
        self.tableWidget.setHorizontalHeaderLabels(header)
        for i, name in enumerate(names):
            self.tableWidget.setRowCount(self.tableWidget.rowCount() + 1)
            for j, title in enumerate(header):
                if j == 0:
                    self.tableWidget.setItem(i, j, QTableWidgetItem(name))
                    continue
                work_id = db_sess.query(Work).filter(Work.title == title).first().id
                pupil_id = db_sess.query(Pupil).filter(Pupil.full_name == name).first().id
                print(title, name)
                mark = db_sess.query(Journal).filter(Journal.task_name == work_id, Journal.full_name == pupil_id).first().mark
                if mark == -1:
                    self.tableWidget.setItem(i, j, QTableWidgetItem('н'))
                else:
                    self.tableWidget.setItem(i, j,  QTableWidgetItem(str(mark)))


class AddTable(QWidget):
    def __init__(self, name):
        super().__init__()
        self.workbook = openpyxl.open(name)
        uic.loadUi('for_table.ui', self)
        self.label_2.hide()
        self.label_3.hide()
        self.lineEdit_2.hide()
        self.lineEdit_3.hide()
        self.pushButton_4.hide()
        self.pushButton_3.hide()
        self.setWindowTitle('Изменение открытой таблицы')
        self.lineEdit.setText(str(self.workbook.sheetnames[0]))
        self.pushButton.clicked.connect(self.save_table)
        self.pushButton_2.clicked.connect(self.close)
        self.show_table()
        self.pushButton_5.clicked.connect(self.add_work)

    def add_work(self):
        file = QFileDialog.getOpenFileName(self, 'Выберите Файл с работой', '', "Text files (*.txt);;Word files (*.docx)")
        shutil.copy2(file[0], 'works/')
        self.file_name = 'works/' + file[0].split('/')[-1]
        self.key = write_key()
        encrypt(self.file_name, self.key)
        self.file_name = self.file_name.split('.')[0] + '.txt'
        print(self.file_name)

    def close(self):
        self.hide()

    def show_table(self):
        worksheet = self.workbook.worksheets[0]
        self.header = [i.value for i in list(worksheet.rows)[0]]
        self.tableWidget.clear()
        self.tableWidget.setColumnCount(len(self.header))
        if 'ФИО' not in self.header or 'Оценка' not in self.header:
            if self.progress.isHidden():
                raise IndexError
            else:
                msg = QMessageBox()
                msg.setIcon(QMessageBox.Critical)
                msg.setText("Error")
                msg.setInformativeText('Неверно заполнена таблица')
                msg.setWindowTitle("Error")
                msg.exec_()
                self.hide()
                return
        self.tableWidget.setRowCount(0)
        self.tableWidget.setHorizontalHeaderLabels(self.header)
        for i in range(worksheet.max_row - 1):
            self.tableWidget.setRowCount(self.tableWidget.rowCount() + 1)
            row = [i.value for i in list(worksheet.rows)[i + 1]]
            for j, elem in enumerate(row):
                self.tableWidget.setItem(i, j, QTableWidgetItem(str(elem)))

    def save_table(self):
        classes = db_sess.query(Class).filter(Class.form == form.comboBox_3.currentText()).first()
        work_title = self.lineEdit.text()
        work = Work(
            title=work_title,
            form=classes.id,
            file_name=self.file_name,
            key=self.key
        )
        db_sess.add(work)
        db_sess.commit()
        for i in range(self.tableWidget.rowCount()):
            scores = []
            variant = 1
            name = ''
            mark = -1
            for j in range(self.tableWidget.rowCount()):
                line = []
                for j in range(len(self.header)):
                    line.append(self.tableWidget.item(i, j).text())
                name = line[self.header.index('ФИО')]
                mark = line[self.header.index("Оценка")]
                variant = line[self.header.index('Вариант')] if 'Вариант' in self.header else 1
                scores = line[self.header.index('Вариант') + 1:] if 'Вариант' in self.header else line[self.header.index("Оценка") + 1:]
                scores = ';'.join(_ for _ in scores)
            scores = ';'.join(_ for _ in scores)
            task_name = db_sess.query(Work).filter(Work.form == classes.id, Work.title == work_title).first().id
            classes = db_sess.query(Class).filter(Class.form == form.comboBox_3.currentText()).first()
            name = db_sess.query(Pupil).filter(Pupil.full_name == name).first()
            if not name:
                res = db_sess.query(Journal).filter(Journal.task_name == work.id).all()
                for _ in res:
                    db_sess.delete(_)
                    db_sess.commit()
                db_sess.delete(work)
                db_sess.commit()
                msg = QMessageBox()
                msg.setIcon(QMessageBox.Critical)
                msg.setText("Error")
                msg.setInformativeText(f'Неверно написано имя: {self.tableWidget.item(i, 0).text()}')
                msg.setWindowTitle("Error")
                msg.exec_()
                return
            max_score = int(self.lineEdit_4.text()) if self.lineEdit_4.text().isdigit() else 100
            pupil = Journal(
                full_name=name.id,
                pupil_form=classes.id,
                task_name=task_name,
                version=variant,
                score_for_task=scores,
                max_score=max_score,
                mark=mark)
            db_sess.add(pupil)
            db_sess.commit()
        self.progress.setText("Таблица успешно сохранена")
        self.progress.setStyleSheet('color: green')
        self.progress.adjustSize()
        form.main_table()


class NewTable(QDialog):
    def __init__(self, *args):
        self.work, self.journal_work = list(), list()
        super().__init__()
        uic.loadUi('for_table.ui', self)
        self.setWindowTitle(args[0])
        self.changing = False
        self.show_table_flag = False
        self.file_name = ''
        self.key = ''
        self.pushButton_3.clicked.connect(self.show_table)
        self.pushButton.clicked.connect(self.save_table)
        self.pushButton_2.clicked.connect(self.close)
        self.pushButton_4.clicked.connect(self.calculate_grade)
        self.pushButton_5.clicked.connect(self.add_work)

    def change_work(self, name):
        self.changing = True
        classes = db_sess.query(Class).filter(Class.form == form.comboBox_3.currentText()).first()
        work = db_sess.query(Work).filter(Work.title == name, Work.form == classes.id).first()
        self.file_name = work.file_name
        self.key = work.key
        journal_work = db_sess.query(Journal).filter(Journal.pupil_form == classes.id, Journal.task_name == work.id).all()
        self.row = 3 + len(journal_work[0].score_for_task.split(';'))
        self.header = ['ФИО', 'Оценка', 'Вариант']
        self.header += [f'Задание {i + 1}' for i in range(len(journal_work[0].score_for_task.split(';')))]
        self.tableWidget.setColumnCount(self.row)
        self.tableWidget.setRowCount(0)
        self.tableWidget.setHorizontalHeaderLabels(self.header)
        self.lineEdit_4.setText(str(journal_work[0].max_score))
        self.lineEdit.setText(work.title)
        self.lineEdit_2.setText(str(journal_work[0].version))
        self.lineEdit_3.setText(str(len(journal_work[0].score_for_task.split(';'))))
        self.var = 1
        print(journal_work)
        for i, row in enumerate(journal_work):
            self.tableWidget.setRowCount(self.tableWidget.rowCount() + 1)
            name = db_sess.query(Pupil).filter(Pupil.id == row.full_name, Pupil.form == classes.id).first().full_name
            self.tableWidget.setItem(i, 0, QTableWidgetItem(name))
            mark = row.mark
            self.tableWidget.setItem(i, 1, QTableWidgetItem(str(mark)))
            variant = row.version
            self.tableWidget.setItem(i, 2, QTableWidgetItem(str(variant)))
            scores = row.score_for_task.split(';')
            for j in range(len(scores)):
                self.tableWidget.setItem(i, j + 3, QTableWidgetItem(str(scores[j])))
        self.work = work
        self.journal_work = journal_work

    def add_work(self):
        file = QFileDialog.getOpenFileName(self, 'Выберите Файл с работой', '', "Text files (*.txt);;Word files (*.docx)")
        shutil.copy2(file[0], 'works/')
        self.file_name = 'works/' + file[0].split('/')[-1]
        self.key = write_key()
        encrypt(self.file_name, self.key)
        self.file_name = self.file_name.split('.')[0] + '.txt'
        print(self.file_name)

    def show_table(self):
        self.tableWidget.clear()
        self.header = ['ФИО', 'Оценка']
        classes = db_sess.query(Class).filter(Class.form == form.comboBox_3.currentText()).first()
        names = db_sess.query(Pupil).filter(Pupil.form == classes.id).all()
        if self.lineEdit_2.text().isdigit():
            self.var = 0 if self.lineEdit_2.text() == '0' or self.lineEdit_2.text() == '1' else 1
        else:
            self.progress.setText('Ошибка: неправильно указано количество вариантов')
            self.progress.setStyleSheet('color: red')
            self.progress.adjustSize()
            return
        if self.lineEdit_3.text().isdigit():
            self.num = int(self.lineEdit_3.text())
        else:
            self.progress.setText('Ошибка: неправильно указано количество заданий')
            self.progress.setStyleSheet('color: red')
            self.progress.adjustSize()
            return
        if self.var:
            self.header.append('Вариант')
        zadaniya = []
        for i in range(1, self.num + 1):
            zadaniya.append(f'Задание⠀{i}')
        self.header += zadaniya
        self.row = self.var + self.num + 2
        self.tableWidget.setColumnCount(self.row)
        self.tableWidget.setRowCount(0)
        self.tableWidget.setHorizontalHeaderLabels(self.header)
        for i, row in enumerate(names):
            self.tableWidget.setRowCount(self.tableWidget.rowCount() + 1)
            for j in range(3):
                if j == 0:
                    self.tableWidget.setItem(i, j, QTableWidgetItem(str(row.full_name)))
        self.show_table_flag = True

    def calculate_grade(self):
        if not self.show_table_flag:
            self.progress.setText('Ошибка: ошибка в заполнении таблицы')
            self.progress.setStyleSheet('color: red')
            self.progress.adjustSize()
            return
        if not self.lineEdit_4.text().isdigit or int(self.lineEdit_4.text()) < 0:
            self.progress.setText('Ошибка: введите корректный максимальный балл')
            self.progress.setStyleSheet('color: red')
            self.progress.adjustSize()
            return
        grade = int(self.lineEdit_4.text())
        for i in range(self.tableWidget.rowCount()):
            n = 0
            start = 3 if self.var else 2
            for j in range(start, self.tableWidget.columnCount()):
                try:
                    n += int(self.tableWidget.item(i, j).text())
                except Exception:
                    mistake(f'Неверно заполнена таблица')
                    return
            if n > grade * 0.85:
                self.tableWidget.setItem(i, 1, QTableWidgetItem(str(5)))
            elif n > grade * 0.65:
                self.tableWidget.setItem(i, 1, QTableWidgetItem(str(4)))
            elif n > grade * 0.4:
                self.tableWidget.setItem(i, 1, QTableWidgetItem(str(3)))
            else:
                self.tableWidget.setItem(i, 1, QTableWidgetItem(str(2)))

    def save_table(self):
        if self.changing:
            print(self.journal_work)
            for i in self.journal_work:
                db_sess.delete(i)
                db_sess.commit()
            db_sess.delete(self.work)
            db_sess.commit()
            self.show_table_flag = True
        if not self.show_table_flag:
            self.progress.setText('Ошибка: ошибка в заполнении таблицы')
            self.progress.setStyleSheet('color: red')
            self.progress.adjustSize()
            return
        classes = db_sess.query(Class).filter(Class.form == form.comboBox_3.currentText()).first()
        work_title = self.lineEdit.text()
        work = Work(
            title=work_title,
            form=classes.id,
            file_name=self.file_name,
            key=self.key
        )
        db_sess.add(work)
        db_sess.commit()
        for i in range(self.tableWidget.rowCount()):
            line = []
            for j in range(len(self.header)):
                try:
                    line.append(self.tableWidget.item(i, j).text())
                except Exception:
                    mistake(f'Неверно заполнена таблица', work)
                    return
            name = line[self.header.index('ФИО')]
            mark = line[self.header.index("Оценка")]
            mark = -1 if mark == 'н' else mark
            variant = line[self.header.index('Вариант')] if 'Вариант' in self.header else 1
            scores = line[self.header.index('Вариант') + 1:] if 'Вариант' in self.header else line[self.header.index("Оценка") + 1:]
            scores = ''.join(i for i in scores)
            if not scores.isdigit():
                db_sess.delete(work)
                db_sess.commit()
                mistake(f'Неверно заполнена таблица')
                return
            scores = ';'.join(_ for _ in scores)
            task_name = db_sess.query(Work).filter(Work.form == classes.id, Work.title == work_title).first().id
            classes = db_sess.query(Class).filter(Class.form == form.comboBox_3.currentText()).first()
            name = db_sess.query(Pupil).filter(Pupil.full_name == name).first()
            if not name:
                db_sess.delete(work)
                db_sess.commit()
                mistake(f'Неверно написано имя: {name}')
                return
            pupil = Journal(
                full_name=name.id,
                pupil_form=classes.id,
                task_name=task_name,
                version=variant,
                score_for_task=scores,
                max_score=int(self.lineEdit_4.text()),
                mark=mark)
            db_sess.add(pupil)
            db_sess.commit()
        self.progress.setText("Таблица успешно сохранена")
        self.progress.setStyleSheet('color: green')
        self.progress.adjustSize()
        form.main_table()

    def close(self):
        self.hide()
        if self.changing:
            self.save_table()


def except_hook(cls, exception, traceback):
    sys.__excepthook__(cls, exception, traceback)


if __name__ == '__main__':
    app = QApplication(sys.argv)
    db_session.global_init('db/journal.db')
    db_sess = db_session.create_session()
    form = MainForm()
    form.show()
    sys.excepthook = except_hook
    sys.exit(app.exec())